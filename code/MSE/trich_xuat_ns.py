import openpyxl  # Thay thế cho pandas
# Xóa các import không cần thiết khác

def main():
    # --- THAY THẾ CÁC THÔNG SỐ CỦA BẠN VÀO ĐÂY ---
    ten_file_excel = r"C:\Users\MSI-PC\Downloads\data Câu 19 (1).xlsx"
    ten_sheet_du_lieu = 'Sheet1'
    ten_tieu_de_cot_x = 'x' # Tên cột 'x' trong file Excel
    ten_tieu_de_cot_y = 'y'    # Tên cột 'y' trong file Excel

    x = []
    y = []

    # --- Đọc file Excel bằng Openpyxl (Gọn nhẹ) ---
    try:
        # Mở file
        workbook = openpyxl.load_workbook(ten_file_excel)
        # Chọn sheet
        sheet = workbook[ten_sheet_du_lieu]

        # Tìm vị trí cột 'x' và 'y' ở hàng đầu tiên (hàng tiêu đề)
        x_col_index = -1
        y_col_index = -1
        for col_idx in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=1, column=col_idx).value
            if cell_value == ten_tieu_de_cot_x:
                x_col_index = col_idx
            if cell_value == ten_tieu_de_cot_y:
                y_col_index = col_idx

        # Kiểm tra xem có tìm thấy cột không
        if x_col_index == -1:
            print(f"Lỗi: Không tìm thấy cột '{ten_tieu_de_cot_x}'")
            return
        if y_col_index == -1:
            print(f"Lỗi: Không tìm thấy cột '{ten_tieu_de_cot_y}'")
            return

        # Đọc dữ liệu từ hàng thứ 2 trở đi (bỏ qua hàng tiêu đề)
        for row_idx in range(2, sheet.max_row + 1):
            x_val = sheet.cell(row=row_idx, column=x_col_index).value
            y_val = sheet.cell(row=row_idx, column=y_col_index).value

            # Đảm bảo dữ liệu không rỗng
            if x_val is not None and y_val is not None:
                x.append(float(x_val))
                y.append(float(y_val))

    except FileNotFoundError:
        print(f"Lỗi: Không tìm thấy file '{ten_file_excel}'.")
        return
    except Exception as e:
        print(f"Đã xảy ra lỗi khi đọc file: {e}")
        return
    # --------------------------------------------------

    # Input: gia tri can tinh noi suy va so diem can trich xuat
    value = float(input("Nhap gia tri can tinh noi suy: "))
    k = int(input("Nhap so diem noi suy can trich xuat: "))

    # Trich xuat diem noi suy (Hàm này giữ nguyên như code của bạn)
    x_new, y_new = trich_xuat_ns(x, y, k, value)

    # In ra man hinh va luu vao file
    print("\nDiem noi suy duoc trich xuat:")
    for i in range(len(x_new)):
        print(f"x[{i}] = {x_new[i]:.7f}, y[{i}] = {y_new[i]:.7f}")
    with open("diem_noi_suy_trich_xuat.txt", "w") as file:
        file.write("Diem noi suy duoc trich xuat:\n")
        for i in range(len(x_new)):
            file.write(f"x[{i}] = {x_new[i]:.7f}, y[{i}] = {y_new[i]:.7f}\n")


def trich_xuat_ns(x, y, k, value):
    # ... (Hàm này của bạn giữ nguyên, không thay đổi)
    N = len(x)
    counter = 0
    if value < x[0] or value > x[N-1]:
        print("Gia tri can tinh noi suy nam ngoai khoang du lieu x.")
        return x, y
    if k >= N:
        print("So diem noi suy lon hon hoac bang so diem co san.")
        return x, y
    left = -1
    right = -1
    for i in range(N - 1):
        if x[i] <= value < x[i+1]:
            left = i
            right = i + 1
            counter = 2
            break
    if value == x[N-1]:
        left = N - 2
        right = N - 1
        counter = 2
    if left == -1:
        print("Không tìm thấy khoảng giá trị phù hợp.")
        return [], []
    while counter < k:
        can_mo_rong_trai = left - 1 >= 0
        can_mo_rong_phai = right + 1 < N
        if not can_mo_rong_trai and not can_mo_rong_phai:
            break
        if can_mo_rong_trai:
            left -= 1
            counter += 1
            if counter == k:
                break
        if can_mo_rong_phai:
            right += 1
            counter += 1
            if counter == k:
                break
    x_new = x[left : right + 1]
    y_new = y[left : right + 1]
    return x_new, y_new


if __name__ == "__main__":
    main()